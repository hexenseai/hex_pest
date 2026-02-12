"""
İstasyon raporu: Tesis + tarih aralığı seçilir, Excel indirilir.
Bölüm (zone) bazında istasyonlar; her iş kaydı tarihi bir sütun, hücrede tüketim (Var/Yok).
"""
from collections import defaultdict
from datetime import date
from io import BytesIO

from django import forms
from django.http import HttpResponse

from django.db.models import Q

from .models import Facility, WorkRecord, WorkRecordStationCount, Station


class IstasyonRaporuForm(forms.Form):
    """Tesis ve tarih aralığı seçimi."""
    tesis = forms.ModelChoiceField(
        label="Tesis",
        queryset=Facility.objects.none(),
        required=True,
    )
    baslangic_tarihi = forms.DateField(
        label="Başlangıç tarihi",
        required=True,
        widget=forms.DateInput(attrs={"type": "date", "class": "w-full rounded border border-base-300 dark:border-base-600 bg-white dark:bg-base-800 px-3 py-2"}),
    )
    bitis_tarihi = forms.DateField(
        label="Bitiş tarihi",
        required=True,
        widget=forms.DateInput(attrs={"type": "date", "class": "w-full rounded border border-base-300 dark:border-base-600 bg-white dark:bg-base-800 px-3 py-2"}),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["tesis"].queryset = Facility.objects.select_related("customer").order_by("customer__kod", "kod")
        self.fields["tesis"].widget.attrs["class"] = "w-full rounded border border-base-300 dark:border-base-600 bg-white dark:bg-base-800 px-3 py-2"

    def clean(self):
        data = super().clean()
        if data.get("baslangic_tarihi") and data.get("bitis_tarihi"):
            if data["baslangic_tarihi"] > data["bitis_tarihi"]:
                raise forms.ValidationError("Bitiş tarihi başlangıçtan önce olamaz.")
        return data


def get_istasyon_raporu_data(facility_id, start_date, end_date):
    """
    Tesis ve tarih aralığına göre rapor verisini döndürür.
    Dönüş: musteri_tesis, adres, date_headers, rows, ratio_genel, ratio_by_zone, zone_stats.
    """
    facility = Facility.objects.select_related("customer").get(pk=facility_id)
    customer = facility.customer
    musteri_tesis = f"{customer.firma_ismi} - {facility.ad}"
    adres = (facility.adres or "").strip() or (customer.adres or "").strip() or "—"

    work_records = list(
        WorkRecord.objects.filter(
            Q(kapatilan_talep__facility_id=facility_id) | Q(facility_id=facility_id),
            tarih__gte=start_date,
            tarih__lte=end_date,
        )
        .order_by("tarih")
        .values_list("id", "tarih")
    )
    wr_ids = [wr[0] for wr in work_records]
    stations = list(
        Station.objects.filter(zone__facility_id=facility_id)
        .select_related("zone")
        .order_by("zone__kod", "kod")
    )
    counts_qs = WorkRecordStationCount.objects.filter(
        work_record_id__in=wr_ids,
        station_id__in=[s.id for s in stations],
    ).values_list("work_record_id", "station_id", "tuketim_var")
    count_map = {(wr_id, st_id): val for wr_id, st_id, val in counts_qs}

    date_headers = [wr[1].strftime("%d.%m.%Y") for wr in work_records]
    rows = []
    for station in stations:
        zone_name = str(station.zone) if station.zone else "—"
        counts = [count_map.get((wr[0], station.id)) for wr in work_records]
        rows.append({
            "zone": zone_name,
            "kod": station.kod,
            "ad": station.ad or "",
            "counts": counts,
        })

    total_stations = len(stations)
    # Tüketim oranı (genel): her tarih için Var/toplam
    ratio_genel = []
    for wr in work_records:
        var_count = sum(1 for i, r in enumerate(rows) if count_map.get((wr[0], stations[i].id)) is True)
        ratio_genel.append(var_count / total_stations if total_stations else 0)

    # Bölge bazlı istasyon grupları
    zone_stations = defaultdict(list)
    for station in stations:
        zone_name = str(station.zone) if station.zone else "—"
        zone_stations[zone_name].append(station)

    # Tüketim oranı (bölge): her bölge için her tarihteki Var/toplam(bölge)
    ratio_by_zone = {}
    for zone_name, st_list in zone_stations.items():
        zone_total = len(st_list)
        ratios = []
        for wr in work_records:
            var_count = sum(1 for st in st_list if count_map.get((wr[0], st.id)) is True)
            ratios.append(var_count / zone_total if zone_total else 0)
        ratio_by_zone[zone_name] = ratios

    # Bölge istatistikleri: tüketim değişim oranı, istasyon değişimi oranı
    zone_stats = []
    if len(wr_ids) >= 2:
        for zone_name in sorted(zone_stations.keys()):
            st_list = zone_stations[zone_name]
            zone_total = len(st_list)
            if zone_total == 0:
                continue
            var_ilk = sum(1 for st in st_list if count_map.get((wr_ids[0], st.id)) is True)
            var_son = sum(1 for st in st_list if count_map.get((wr_ids[-1], st.id)) is True)

            # Tüketim değişim oranı: (Var_son - Var_ilk) / Var_ilk * 100 (Var_ilk > 0 ise)
            if var_ilk > 0:
                tuketim_degisim = ((var_son - var_ilk) / var_ilk) * 100
            else:
                tuketim_degisim = (var_son * 100) if var_son > 0 else 0  # 0'dan artış

            # İstasyon değişimi: değişen istasyon sayısı (tarihler arası Var<->Yok değişen) / zone_total * 100
            degisen = 0
            for st in st_list:
                val_ilk = count_map.get((wr_ids[0], st.id)) is True
                val_son = count_map.get((wr_ids[-1], st.id)) is True
                if val_ilk != val_son:
                    degisen += 1
            istasyon_degisim = (degisen / zone_total * 100) if zone_total else 0

            zone_stats.append({
                "zone": zone_name,
                "tuketim_degisim": tuketim_degisim,
                "istasyon_degisim": istasyon_degisim,
                "var_ilk": var_ilk,
                "var_son": var_son,
                "zone_total": zone_total,
            })

    return {
        "musteri_tesis": musteri_tesis,
        "adres": adres,
        "date_headers": date_headers,
        "rows": rows,
        "ratio_genel": ratio_genel,
        "ratio_by_zone": ratio_by_zone,
        "zone_stats": zone_stats,
    }


def _format_percent(val):
    """Oranı % formatında yaz. val 0-1 aralığında veya zaten yüzde (örn. 40) olabilir."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        pct = val * 100 if val <= 1 else val
        return f"{pct:.1f}%"
    return str(val)


def build_istasyon_raporu_excel(facility_id, start_date, end_date):
    """
    Tesis ve tarih aralığına göre Excel dosyası üretir.
    İlk iki satır: müşteri adı - tesis, alt satırda adres.
    Başlık satırı altında: tüketim oranı (genel), bölge bazlı oranlar.
    Sonra satırlar: bölüm bazında istasyonlar (Bölge, İstasyon Kodu, Açıklama).
    Sütunlar: her iş kaydı tarihi için tüketim (Var/Yok).
    En sonda: bölge istatistikleri (tüketim değişim oranı, istasyon değişimi oranı).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    data = get_istasyon_raporu_data(facility_id, start_date, end_date)
    musteri_tesis = data["musteri_tesis"]
    adres = data["adres"]
    date_headers = data["date_headers"]
    rows = data["rows"]
    ratio_genel = data.get("ratio_genel", [])
    ratio_by_zone = data.get("ratio_by_zone", {})
    zone_stats = data.get("zone_stats", [])
    ncols = 3 + len(date_headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İstasyon Raporu"

    thin = Side(style="thin")
    header_font = Font(bold=True)
    small_font = Font(size=9)

    # İlk iki satır: müşteri - tesis, adres
    ws.cell(row=1, column=1, value=musteri_tesis)
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(3, ncols))
    ws.row_dimensions[1].height = 20
    ws.cell(row=2, column=1, value=adres)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(3, ncols))
    ws.row_dimensions[2].height = 18

    # Başlık satırı (3): Bölge | İstasyon Kodu | Açıklama | Tarih1 | Tarih2 | ...
    header_row = 3
    headers = ["Bölge", "İstasyon Kodu", "Açıklama"] + date_headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[header_row].height = 22

    # Tarih sütunları altına: tüketim oranı (genel)
    current_row = header_row + 1
    if ratio_genel:
        ws.cell(row=current_row, column=1, value="Tüketim oranı (genel)")
        ws.cell(row=current_row, column=1).font = small_font
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        for col_idx, val in enumerate(ratio_genel, 4):
            cell = ws.cell(row=current_row, column=col_idx, value=_format_percent(val))
            cell.font = small_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        current_row += 1

        # Bölge bazlı oranlar
        for zone_name in sorted(ratio_by_zone.keys()):
            ratios = ratio_by_zone[zone_name]
            ws.cell(row=current_row, column=1, value=f"Tüketim oranı - {zone_name}")
            ws.cell(row=current_row, column=1).font = small_font
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            for col_idx, val in enumerate(ratios, 4):
                cell = ws.cell(row=current_row, column=col_idx, value=_format_percent(val))
                cell.font = small_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            current_row += 1

        # Boş satır ayırıcı
        current_row += 1

    # İstasyon satırları
    for row in rows:
        ws.cell(row=current_row, column=1, value=row["zone"])
        ws.cell(row=current_row, column=2, value=row["kod"])
        ws.cell(row=current_row, column=3, value=row["ad"])
        for col_idx, val in enumerate(row["counts"], 4):
            cell_val = "Var" if val else ("Yok" if val is False else "")
            cell = ws.cell(row=current_row, column=col_idx, value=cell_val)
            cell.alignment = Alignment(horizontal="center" if cell_val else "left", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Bölge istatistikleri (tüketim değişim oranı, istasyon değişimi oranı)
    current_row += 1
    if zone_stats:
        ws.cell(row=current_row, column=1, value="Bölge İstatistikleri")
        ws.cell(row=current_row, column=1).font = header_font
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        stat_headers = ["Bölge", "Tüketim değişim oranı", "İstasyon değişimi oranı", "Var (ilk)", "Var (son)", "Toplam istasyon"]
        for col, h in enumerate(stat_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        current_row += 1

        for stat in zone_stats:
            ws.cell(row=current_row, column=1, value=stat["zone"])
            ws.cell(row=current_row, column=2, value=_format_percent(stat["tuketim_degisim"]))
            ws.cell(row=current_row, column=3, value=_format_percent(stat["istasyon_degisim"]))
            ws.cell(row=current_row, column=4, value=stat["var_ilk"])
            ws.cell(row=current_row, column=5, value=stat["var_son"])
            ws.cell(row=current_row, column=6, value=stat["zone_total"])
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            current_row += 1

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    for col_idx in range(4, 4 + len(date_headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
